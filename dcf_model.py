"""
DCF Valuation Model
====================
Pulls 5 years of historical financials, projects FCF, computes WACC,
runs sensitivity analysis, and exports a formatted Excel workbook.

Stack: Python, yfinance, pandas, NumPy, openpyxl
"""

from __future__ import annotations

#imports to work with command line
import sys
import warnings
from dataclasses import dataclass, field
from typing import Optional

#imports for math and excel functionality
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

warnings.filterwarnings("ignore")

# ----------------------------------------------------------------------------
# CONFIGURATION
# ----------------------------------------------------------------------------

#default company and years
TICKER = "AAPL"
PROJECTION_YEARS = 5
HISTORICAL_YEARS = 5
OUTPUT_FILE = "DCF_Valuation.xlsx"

# Macroeconomic assumptions (override after pulling data)

#Approximate 10Y US Treasury yield
RISK_FREE_RATE = 0.043
#Damodaran implied ERP
EQUITY_RISK_PREMIUM = 0.055
#Approximate long-run nominal GDP
TERMINAL_GROWTH = 0.025

# ----------------------------------------------------------------------------
# DATA STRUCTURES
# ----------------------------------------------------------------------------

#dataclass to hold all of our required information on ticker's company
@dataclass
class CompanyData:
    ticker: str
    name: str = ""
    sector: str = ""
    currency: str = "USD"
    current_price: float = 0.0

    #in millions
    shares_outstanding: float = 0.0
    market_cap: float = 0.0
    total_debt: float = 0.0
    cash_and_equivalents: float = 0.0
    interest_expense: float = 0.0

    beta: float = 1.0
    effective_tax_rate: float = 0.21
    fiscal_years: list = field(default_factory=list)
    revenue: list = field(default_factory=list)
    ebit: list = field(default_factory=list)
    #depreciation & amortization
    da: list = field(default_factory=list)
    capex: list = field(default_factory=list)
    working_capital: list = field(default_factory=list)
    data_source: str = "live"

    #helper function to calculate total debt - cash
    @property
    def net_debt(self) -> float:
        return self.total_debt - self.cash_and_equivalents

    #helper function to calculate pretax cost of debt
    @property
    def pretax_cost_of_debt(self) -> float:
        #If no debt return 5%
        if self.total_debt is None or self.total_debt <= 0:
            return 0.05
        ie = self.interest_expense
        # if it is missing return 4% as a placeholder, sensible default for investment-grade issuers
        if ie is None or (isinstance(ie, float) and np.isnan(ie)) or ie <= 0:
            return 0.04
        #else, calculate interest expense / total debt and clamp to 1% to 15%
        #clamp to guard against weird values like 0.0000001% arising
        return max(min(ie / self.total_debt, 0.15), 0.01)


# ----------------------------------------------------------------------------
# DATA FETCHING
# ----------------------------------------------------------------------------
def fetch_data_yfinance(ticker: str) -> CompanyData:
    #Pull live data from yfinance. Returns CompanyData or raises on failure
    import yfinance as yf

    #handle to company's data, .info retrieves dictionary of summary metrics
    yf_ticker = yf.Ticker(ticker)
    info = yf_ticker.info
    if not info or "currentPrice" not in info:
        raise RuntimeError("yfinance returned no usable data")

    #retrieve metrics
    cd = CompanyData(ticker=ticker)
    cd.name = info.get("longName", ticker)
    cd.sector = info.get("sector", "")
    cd.currency = info.get("financialCurrency", "USD")
    cd.current_price = float(info.get("currentPrice") or info.get("regularMarketPrice") or 0)
    cd.shares_outstanding = float(info.get("sharesOutstanding", 0) or 0) / 1e6
    cd.market_cap = float(info.get("marketCap", 0) or 0) / 1e6
    cd.beta = float(info.get("beta") or 1.0)

    #retrieve income statement as pandas data frame
    fin = yf_ticker.financials
    if fin is None or fin.empty:
        raise RuntimeError("Income statement unavailable")
    #order columns from oldest -> newest
    fin = fin.iloc[:, ::-1]

    #helper function to search multiple possible names for same line item and return
    #(sometimes yahoo field names are different between versions/companies)
    def first_match(df, keys):
        for k in keys:
            if k in df.index:
                return df.loc[k]
        return None

    #each locates one row of the income statement
    revenue_row = first_match(fin, ["Total Revenue", "TotalRevenue", "Revenue"])
    ebit_row = first_match(fin, ["Operating Income", "EBIT", "OperatingIncome"])
    interest_row = first_match(fin, [
        "Interest Expense",
        "InterestExpense",
        "Interest Expense Non Operating",
        "InterestExpenseNonOperating",
        "Net Non Operating Interest Income Expense",
    ])
    tax_row = first_match(fin, ["Tax Provision", "Income Tax Expense", "TaxProvision"])
    pretax_row = first_match(fin, ["Pretax Income", "Income Before Tax", "PretaxIncome"])

    #fail if we are missing revenue or ebit
    if revenue_row is None or ebit_row is None:
        raise RuntimeError("Required income-statement lines missing")
    
    #retrieve list of fiscal years, convert dollar values (revenue & EBIT) to millions
    cd.fiscal_years = [c.strftime("FY%Y") for c in revenue_row.index]
    cd.revenue = (revenue_row.values / 1e6).tolist()
    cd.ebit = (ebit_row.values / 1e6).tolist()

    #Drop any years where revenue is missing (yfinance sometimes returns NaN for the oldest year).
    valid = [i for i, v in enumerate(cd.revenue) if pd.notna(v) and v > 0]
    #use this to align the cash-flow rows below
    valid_dates = revenue_row.index[valid]
    cd.fiscal_years = [cd.fiscal_years[i] for i in valid]
    cd.revenue = [cd.revenue[i] for i in valid]
    cd.ebit = [cd.ebit[i] for i in valid]

    #cash flow statement
    cf = yf_ticker.cashflow
    if cf is not None and not cf.empty:
        cf = cf.iloc[:, ::-1]
        #find depreciation and amortization
        da_row = first_match(cf, [
            "Depreciation And Amortization",
            "Depreciation Amortization Depletion",
            "DepreciationAndAmortization",
            "Depreciation",
        ])
        #find capital expenditure
        capex_row = first_match(cf, [
            "Capital Expenditure",
            "CapitalExpenditure",
            "Capital Expenditures",
        ])
        #find change in working capital
        wc_row = first_match(cf, [
            "Change In Working Capital",
            "ChangeInWorkingCapital",
        ])
        #convert existing rows to millions
        if da_row is not None:
            cd.da = (da_row.reindex(valid_dates).fillna(0).values / 1e6).tolist()
        if capex_row is not None:
            #yfinance reports capex as negative; we want it as a positive number to subtract
            cd.capex = (capex_row.reindex(valid_dates).abs().fillna(0).values / 1e6).tolist()
        if wc_row is not None:
            cd.working_capital = (wc_row.reindex(valid_dates).fillna(0).values / 1e6).tolist()

    #pull the most recent column from the balance sheet
    bs = yf_ticker.balance_sheet
    if bs is not None and not bs.empty:
        most_recent = bs.iloc[:, 0]
        cd.total_debt = float(most_recent.get("Total Debt", 0) or 0) / 1e6
        cash = (most_recent.get("Cash And Cash Equivalents", 0)
                or most_recent.get("Cash", 0) or 0)
        sti = most_recent.get("Other Short Term Investments", 0) or 0
        cd.cash_and_equivalents = (float(cash) + float(sti)) / 1e6

    #retrieve most recent interest expense (skip NaN / zero values)
    if interest_row is not None and len(interest_row) > 0:
        for v in reversed(list(interest_row.values)):
            if pd.notna(v) and float(v) != 0:
                cd.interest_expense = abs(float(v)) / 1e6
                break

    #retrieve the effective tax rate as a multi-year average
    if tax_row is not None and pretax_row is not None:
        taxes = tax_row.fillna(0).values
        pretax = pretax_row.fillna(0).values
        if pretax.sum() > 0:
            cd.effective_tax_rate = float(np.clip(taxes.sum() / pretax.sum(), 0.10, 0.35))


    #use only the last n years
    n = HISTORICAL_YEARS
    cd.fiscal_years = cd.fiscal_years[-n:]
    cd.revenue = cd.revenue[-n:]
    cd.ebit = cd.ebit[-n:]
    if cd.da:
        cd.da = cd.da[-n:]
    if cd.capex:
        cd.capex = cd.capex[-n:]
    if cd.working_capital:
        cd.working_capital = cd.working_capital[-n:]

    cd.data_source = "yfinance (live)"
    return cd

#safety net against blank cells from NaN returns
def _sanitize(cd: CompanyData) -> CompanyData:
    #Replace any NaN/None scalars with sensible defaults so Excel never sees blanks
    def clean(v, default=0.0):
        if v is None:
            return default
        try:
            if isinstance(v, float) and np.isnan(v):
                return default
        except Exception:
            pass
        return v

    #clean retrieved metrics
    cd.current_price = float(clean(cd.current_price))
    cd.shares_outstanding = float(clean(cd.shares_outstanding))
    cd.market_cap = float(clean(cd.market_cap))
    cd.total_debt = float(clean(cd.total_debt))
    cd.cash_and_equivalents = float(clean(cd.cash_and_equivalents))
    cd.beta = float(clean(cd.beta, default=1.0))
    cd.interest_expense = float(clean(cd.interest_expense))
    cd.effective_tax_rate = float(clean(cd.effective_tax_rate, default=0.21))

    #helper function like prior, but now to clean lists
    def clean_list(lst, default=0.0):
        return [float(clean(x, default)) for x in lst]

    cd.revenue = clean_list(cd.revenue)
    cd.ebit = clean_list(cd.ebit)
    cd.da = clean_list(cd.da)
    cd.capex = clean_list(cd.capex)
    cd.working_capital = clean_list(cd.working_capital)
    return cd

#returns cleaned fetched data
def fetch_company_data(ticker: str) -> CompanyData:
    #pull live data from yfinance and sanitize it for Excel output
    print(f"[1/4] Fetching live data for {ticker} via yfinance...")
    try:
        cd = fetch_data_yfinance(ticker)
    except Exception as e:
        print(f"\nERROR: Could not fetch data for {ticker!r}.")
        print(f"       Reason: {type(e).__name__}: {e}")
        print(f"\n       Possible causes:")
        print(f"         - No internet connection")
        print(f"         - Yahoo Finance is rate-limiting requests (try again in a minute)")
        print(f"         - Ticker symbol is invalid or delisted")
        sys.exit(1)
    print(f"      OK - {cd.name}, {len(cd.fiscal_years)} years of history")
    return _sanitize(cd)


# ----------------------------------------------------------------------------
# STYLING HELPERS (openpyxl)
# ----------------------------------------------------------------------------

#helpers to make xlsx file look professional
FONT_NAME = "Arial"
#dark blue
HEADER_FILL = PatternFill("solid", start_color="1F3864")
SUBHEADER_FILL = PatternFill("solid", start_color="D9E1F2")
#yellow for key inputs
INPUT_FILL = PatternFill("solid", start_color="FFFF00")
#light green
RESULT_FILL = PatternFill("solid", start_color="E2EFDA")
BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

FMT_CURRENCY = '"$"#,##0;("$"#,##0);"-"'
#put negative numbers and 0 (-) in correct accounting convention
FMT_CURRENCY_DEC = '"$"#,##0.00;("$"#,##0.00);"-"'
FMT_NUM = '#,##0;(#,##0);"-"'
FMT_PCT = "0.0%;(0.0%);-"
FMT_MULTIPLE = "0.00x"
FMT_YEAR = "@"

#helper function to style worksheet headers
def style_header(cell, large=False):
    cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF",
                     size=14 if large else 11)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left" if large else "center",
                               vertical="center")

#similar to the last helper but different conventions for sub headers
def style_subheader(cell):
    cell.font = Font(name=FONT_NAME, bold=True, color="000000")
    cell.fill = SUBHEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER_THIN

#styler for input cells
def style_input(cell, fmt=None):
    # blue = hardcoded input
    cell.font = Font(name=FONT_NAME, color="0000FF")
    cell.fill = INPUT_FILL
    if fmt:
        cell.number_format = fmt
    cell.border = BORDER_THIN
    cell.alignment = Alignment(horizontal="right")

#styler for formula cells
def style_formula(cell, fmt=None, link=False):
    cell.font = Font(name=FONT_NAME, color="008000" if link else "000000")
    if fmt:
        cell.number_format = fmt
    cell.border = BORDER_THIN
    cell.alignment = Alignment(horizontal="right")

#styler for labels
def style_label(cell, bold=False, indent=0):
    cell.font = Font(name=FONT_NAME, bold=bold, color="000000")
    cell.alignment = Alignment(horizontal="left", indent=indent)
    cell.border = BORDER_THIN

#styler for results
def style_result(cell, fmt=None):
    cell.font = Font(name=FONT_NAME, bold=True, color="000000")
    cell.fill = RESULT_FILL
    if fmt:
        cell.number_format = fmt
    cell.border = BORDER_THIN
    cell.alignment = Alignment(horizontal="right")

#helper to format column widths to fit items
def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ----------------------------------------------------------------------------
# WORKBOOK BUILDERS
# ----------------------------------------------------------------------------

#ASSUMPTIONS SHEET

#first of five sheet makers
#take company data and create assumption sheet via a refs dict
#the refs dictionary is used to write cross sheet formulas
def build_assumptions_sheet(wb: Workbook, cd: CompanyData) -> dict:
    #inputs sheet. Returns a dict of named cell references for cross-sheet links
    ws = wb.create_sheet("Assumptions")
    refs = {}

    #title
    ws.merge_cells("A1:C1")
    ws["A1"] = f"Assumptions - {cd.name} ({cd.ticker})"
    style_header(ws["A1"], large=True)
    ws.row_dimensions[1].height = 26

    # ---- Company Data ----
    
    #store the ticker at cell A3
    row = 3
    ws[f"A{row}"] = "Company Info"
    style_subheader(ws[f"A{row}"])
    ws.merge_cells(f"A{row}:C{row}")
    row += 1
    #initialize a list of three element tuples of form (label, value, format)
    company_info = [
        ("Ticker", cd.ticker, "@"),
        ("Company Name", cd.name, "@"),
        ("Sector", cd.sector, "@"),
        ("Reporting Currency", cd.currency, "@"),
        ("Data Source", cd.data_source, "@"),
    ]
    #iterate through the company info, at each element pull the label, value and format of our respective tuple in list
    for label, val, fmt in company_info:
        
        #give our row the correct label & style
        ws[f"A{row}"] = label
        style_label(ws[f"A{row}"])

        #give corressponding value and format it
        ws[f"B{row}"] = val
        ws[f"B{row}"].font = Font(name=FONT_NAME, color="000000")
        ws[f"B{row}"].number_format = fmt
        ws[f"B{row}"].border = BORDER_THIN
        row += 1

    # ---- Market data ----
    row += 1
    ws[f"A{row}"] = "Market Data (most recent)"
    style_subheader(ws[f"A{row}"])
    ws.merge_cells(f"A{row}:C{row}")
    row += 1

    #data block to similar format containing most recent data
    market_data = [
        ("Current Share Price", cd.current_price, FMT_CURRENCY_DEC, "current_price"),
        ("Shares Outstanding (mm)", cd.shares_outstanding, FMT_NUM, "shares_out"),
        ("Market Cap ($mm)", None, FMT_CURRENCY, "market_cap"),  # formula
        ("Total Debt ($mm)", cd.total_debt, FMT_CURRENCY, "total_debt"),
        ("Cash & ST Investments ($mm)", cd.cash_and_equivalents, FMT_CURRENCY, "cash"),
        ("Net Debt ($mm)", None, FMT_CURRENCY, "net_debt"),  # formula
    ]

    #for loop to print metric, and formatted value
    for label, val, fmt, key in market_data:
        ws[f"A{row}"] = label
        style_label(ws[f"A{row}"])
        if val is None:
            #keys determine whether to hadrcode value or a formula
            if key == "market_cap":
                ws[f"B{row}"] = f"=B{refs['_current_price_row']}*B{refs['_shares_out_row']}"
            elif key == "net_debt":
                ws[f"B{row}"] = f"=B{refs['_total_debt_row']}-B{refs['_cash_row']}"
            style_formula(ws[f"B{row}"], fmt)
        else:
            ws[f"B{row}"] = val
            style_input(ws[f"B{row}"], fmt)
        refs[key] = f"Assumptions!$B${row}"
        refs[f"_{key}_row"] = row
        row += 1

    # ---- Cost of capital inputs ----

    #store capital inputs in data block and paste values similarly
    row += 1
    ws[f"A{row}"] = "Cost of Capital Inputs"
    style_subheader(ws[f"A{row}"])
    ws.merge_cells(f"A{row}:C{row}")
    row += 1
    cap_inputs = [
        ("Risk-Free Rate (10Y UST)", RISK_FREE_RATE, FMT_PCT, "rf"),
        ("Equity Risk Premium", EQUITY_RISK_PREMIUM, FMT_PCT, "erp"),
        ("Beta (levered)", cd.beta, "0.00", "beta"),
        ("Pre-Tax Cost of Debt", cd.pretax_cost_of_debt, FMT_PCT, "kd_pretax"),
        ("Effective Tax Rate", cd.effective_tax_rate, FMT_PCT, "tax_rate"),
    ]
    for label, val, fmt, key in cap_inputs:
        ws[f"A{row}"] = label
        style_label(ws[f"A{row}"])
        ws[f"B{row}"] = val
        style_input(ws[f"B{row}"], fmt)
        refs[key] = f"Assumptions!$B${row}"
        refs[f"_{key}_row"] = row
        #source notes
        if key == "kd_pretax":
            ws[f"C{row}"] = f"Interest Exp / Total Debt = {cd.interest_expense:,.0f} / {cd.total_debt:,.0f}"
            ws[f"C{row}"].font = Font(name=FONT_NAME, italic=True, color="595959", size=9)
        row += 1

    # ---- Projection assumptions ----
    row += 1
    ws[f"A{row}"] = "Projection Assumptions (Years 1-5)"
    style_subheader(ws[f"A{row}"])
    ws.merge_cells(f"A{row}:G{row}")
    row += 1

    #Header row for projection years
    ws[f"A{row}"] = "Driver"
    style_subheader(ws[f"A{row}"])
    for i in range(PROJECTION_YEARS):
        col = get_column_letter(2 + i)
        ws[f"{col}{row}"] = f"Year {i+1}"
        style_subheader(ws[f"{col}{row}"])
    row += 1

    #Assume gentle growth deceleration, stable margins for a company like AAPL
    #Note: must change defaults for unconventional large cap companies (NVDA, PLTR)
    growth_defaults = [0.05, 0.05, 0.04, 0.04, 0.03]
    margin_defaults = _historical_margin_path(cd)
    capex_defaults = _historical_ratio_path(cd.capex, cd.revenue, default=0.03)
    da_defaults = _historical_ratio_path(cd.da, cd.revenue, default=0.03)
    #change of NWC as % of incremental revenue
    wc_defaults = [0.005] * PROJECTION_YEARS

    drivers = [
        ("Revenue Growth Rate", growth_defaults, FMT_PCT, "growth"),
        ("EBIT Margin", margin_defaults, FMT_PCT, "ebit_margin"),
        ("D&A as % of Revenue", da_defaults, FMT_PCT, "da_pct"),
        ("CapEx as % of Revenue", capex_defaults, FMT_PCT, "capex_pct"),
        ("ΔNWC as % of Δ Revenue", wc_defaults, FMT_PCT, "wc_pct"),
    ]

    #format 5x5 grid as yellow input, paste formatted input from drivers
    refs["proj_rows"] = {}
    for label, vals, fmt, key in drivers:
        ws[f"A{row}"] = label
        style_label(ws[f"A{row}"])
        for i, v in enumerate(vals):
            col = get_column_letter(2 + i)
            ws[f"{col}{row}"] = v
            style_input(ws[f"{col}{row}"], fmt)
        refs["proj_rows"][key] = row
        row += 1

    # ---- Terminal value inputs ----

    #paste terminal growth rate g from the refs dictionary
    row += 1
    ws[f"A{row}"] = "Terminal Value Inputs"
    style_subheader(ws[f"A{row}"])
    ws.merge_cells(f"A{row}:C{row}")
    row += 1
    ws[f"A{row}"] = "Terminal Growth Rate (g)"
    style_label(ws[f"A{row}"])
    ws[f"B{row}"] = TERMINAL_GROWTH
    style_input(ws[f"B{row}"], FMT_PCT)
    refs["g"] = f"Assumptions!$B${row}"
    refs["_g_row"] = row

    set_col_widths(ws, {"A": 32, "B": 16, "C": 16, "D": 14, "E": 14, "F": 14, "G": 14})
    return refs

#Returns 5 forward EBIT margin assumptions, anchored on recent history
#Note: currently all values are uniform, change in later updates
def _historical_margin_path(cd: CompanyData) -> list:
    #if ebit or revenue is empty return a list of 25% margin
    if not cd.ebit or not cd.revenue:
        return [0.25] * PROJECTION_YEARS
    #take last 3 elements of ebit and revenue then pair them up
    #then divide each ebit by its matching revenue, average them and convert to float
    recent_margin = float(np.mean([e / r for e, r in zip(cd.ebit[-3:], cd.revenue[-3:])]))
    #clamp recent margin between 10% and 35%
    target = max(min(recent_margin, 0.35), 0.10)
    path = []
    for i in range(PROJECTION_YEARS):
        #very mild fade over 5 years
        path.append(round(target - (target - max(target * 0.95, 0.10)) * (i / (PROJECTION_YEARS - 1)), 4))
    return path

#used to set capex and D&A as % of revenue from recent historical averages
#Note: currently all values are constant, change in later updates
def _historical_ratio_path(num: list, den: list, default: float) -> list:
    #if data is missing return the default parameter
    if not num or not den:
        return [default] * PROJECTION_YEARS
    #create a list of 3 ratios from numerator and denominator lists assuming denominator exists
    pairs = [n / d for n, d in zip(num[-3:], den[-3:]) if d]
    if not pairs:
        return [default] * PROJECTION_YEARS
    #average the three ratios and clamp between 0.5% and 20%
    avg = float(np.mean(pairs))
    avg = max(min(avg, 0.20), 0.005)
    return [round(avg, 4)] * PROJECTION_YEARS

#build the Weighted Average Cost of Capital sheet
def build_wacc_sheet(wb: Workbook, cd: CompanyData, refs: dict):
    ws = wb.create_sheet("WACC")
    ws.merge_cells("A1:C1")
    ws["A1"] = "Weighted Average Cost of Capital"
    style_header(ws["A1"], large=True)
    ws.row_dimensions[1].height = 26

    #initialize rows with label, value, format and kind
    rows = []

    #Cost of Equity (CAPM)
    rows.append(("Cost of Equity (CAPM)", None, None, "header"))
    rows.append(("Risk-Free Rate", f"={refs['rf']}", FMT_PCT, "link"))
    rows.append(("Equity Risk Premium", f"={refs['erp']}", FMT_PCT, "link"))
    rows.append(("Beta", f"={refs['beta']}", "0.00", "link"))
    rows.append(("Cost of Equity = Rf + β × ERP", None, FMT_PCT, "ke"))
    rows.append(("", None, None, "blank"))
    #Cost of Debt
    rows.append(("Cost of Debt", None, None, "header"))
    rows.append(("Pre-Tax Cost of Debt", f"={refs['kd_pretax']}", FMT_PCT, "link"))
    rows.append(("Tax Rate", f"={refs['tax_rate']}", FMT_PCT, "link"))
    rows.append(("After-Tax Cost of Debt = Kd × (1 - t)", None, FMT_PCT, "kd"))
    rows.append(("", None, None, "blank"))
    #Capital Structure
    rows.append(("Capital Structure", None, None, "header"))
    rows.append(("Market Cap ($mm)", f"={refs['market_cap']}", FMT_CURRENCY, "link"))
    rows.append(("Total Debt ($mm)", f"={refs['total_debt']}", FMT_CURRENCY, "link"))
    rows.append(("Total Capital (E + D)", None, FMT_CURRENCY, "ev"))
    rows.append(("Weight of Equity (E/V)", None, FMT_PCT, "we"))
    rows.append(("Weight of Debt (D/V)", None, FMT_PCT, "wd"))
    rows.append(("", None, None, "blank"))
    #WACC
    rows.append(("WACC = E/V × Ke + D/V × Kd(1-t)", None, FMT_PCT, "wacc"))

    row_map = {}
    #start at row three to bypass title banner
    for i, (label, value, fmt, kind) in enumerate(rows, start=3):
        #if the row is a header, merge columns and format as header
        ws[f"A{i}"] = label
        if kind == "header":
            style_subheader(ws[f"A{i}"])
            ws.merge_cells(f"A{i}:C{i}")
            continue
        if kind == "blank":
            continue

        style_label(ws[f"A{i}"], indent=1)
        #style cell according to kind and paste value
        if kind == "link":
            ws[f"B{i}"] = value
            style_formula(ws[f"B{i}"], fmt, link=True)
        elif kind == "ke":
            ke_row = i
            row_map["ke"] = ke_row
            # Find rf, erp, beta rows above
            ws[f"B{i}"] = f"=B{row_map['rf']}+B{row_map['beta']}*B{row_map['erp']}"
            style_result(ws[f"B{i}"], fmt)
        elif kind == "kd":
            row_map["kd"] = i
            ws[f"B{i}"] = f"=B{row_map['kd_pretax']}*(1-B{row_map['tax_rate']})"
            style_result(ws[f"B{i}"], fmt)
        elif kind == "ev":
            row_map["ev"] = i
            ws[f"B{i}"] = f"=B{row_map['mcap']}+B{row_map['debt']}"
            style_result(ws[f"B{i}"], fmt)
        elif kind == "we":
            row_map["we"] = i
            ws[f"B{i}"] = f"=B{row_map['mcap']}/B{row_map['ev']}"
            style_result(ws[f"B{i}"], fmt)
        elif kind == "wd":
            row_map["wd"] = i
            ws[f"B{i}"] = f"=B{row_map['debt']}/B{row_map['ev']}"
            style_result(ws[f"B{i}"], fmt)
        #final weighted average cost of capital is
        # (E/V) * Cost of Equity + (D/V) * After-tax Cost of Debt
        elif kind == "wacc":
            row_map["wacc"] = i
            ws[f"A{i}"].font = Font(name=FONT_NAME, bold=True)
            ws[f"B{i}"] = f"=B{row_map['we']}*B{row_map['ke']}+B{row_map['wd']}*B{row_map['kd']}"
            style_result(ws[f"B{i}"], fmt)

        #save row positions for label based references
        if "Risk-Free Rate" in label:
            row_map["rf"] = i
        elif "Equity Risk Premium" in label:
            row_map["erp"] = i
        elif label == "Beta":
            row_map["beta"] = i
        elif "Pre-Tax Cost of Debt" in label:
            row_map["kd_pretax"] = i
        elif label == "Tax Rate":
            row_map["tax_rate"] = i
        elif "Market Cap" in label:
            row_map["mcap"] = i
        elif "Total Debt" in label:
            row_map["debt"] = i

    #save the wacc reference
    refs["wacc"] = f"WACC!$B${row_map['wacc']}"
    refs["_wacc_row"] = row_map["wacc"]

    set_col_widths(ws, {"A": 38, "B": 18, "C": 28})
    return refs


def build_fcf_sheet(wb: Workbook, cd: CompanyData, refs: dict):
    
    #create a sheet for free cash flow projection
    ws = wb.create_sheet("FCF Projections")
    ws.merge_cells("A1:L1")
    ws["A1"] = f"Free Cash Flow - Historical & Projected ($mm)"
    style_header(ws["A1"], large=True)
    ws.row_dimensions[1].height = 26

    #Build year headers
    header_row = 3
    ws[f"A{header_row}"] = "Line Item"
    style_subheader(ws[f"A{header_row}"])

    #get years for future projection and amount of historical years we pulled
    n_hist = len(cd.fiscal_years)
    n_proj = PROJECTION_YEARS

    #Historical columns
    hist_cols = []
    for i, fy in enumerate(cd.fiscal_years):
        col = get_column_letter(2 + i)
        ws[f"{col}{header_row}"] = fy
        style_subheader(ws[f"{col}{header_row}"])
        hist_cols.append(col)

    #projected columns
    proj_cols = []
    for i in range(n_proj):
        col = get_column_letter(2 + n_hist + i)
        ws[f"{col}{header_row}"] = f"Year +{i+1}"
        style_subheader(ws[f"{col}{header_row}"])
        proj_cols.append(col)

    # ---- Historicals ----

    #all of the following metrics are for the historical years after the fact
    r = header_row + 1
    ws[f"A{r}"] = "HISTORICALS"
    ws[f"A{r}"].font = Font(name=FONT_NAME, bold=True)
    r += 1

    #paste revenue for all pulled years
    rev_row = r
    ws[f"A{r}"] = "Revenue"
    style_label(ws[f"A{r}"], indent=1)
    for i, col in enumerate(hist_cols):
        ws[f"{col}{r}"] = cd.revenue[i]
        style_input(ws[f"{col}{r}"], FMT_CURRENCY)
    r += 1

    #compute the growth in revenue from the previous year
    rev_growth_row = r
    ws[f"A{r}"] = "  Revenue Growth"
    style_label(ws[f"A{r}"], indent=1)
    ws[f"{hist_cols[0]}{r}"] = "n/a"
    ws[f"{hist_cols[0]}{r}"].alignment = Alignment(horizontal="right")
    ws[f"{hist_cols[0]}{r}"].font = Font(name=FONT_NAME, color="595959", italic=True)
    for i in range(1, n_hist):
        ws[f"{hist_cols[i]}{r}"] = (
            f"=({hist_cols[i]}{rev_row}-{hist_cols[i-1]}{rev_row})/{hist_cols[i-1]}{rev_row}"
        )
        style_formula(ws[f"{hist_cols[i]}{r}"], FMT_PCT)
    r += 1

    #fill each earnings before interest with corresponding value in cd
    ebit_row = r
    ws[f"A{r}"] = "EBIT (Operating Income)"
    style_label(ws[f"A{r}"], indent=1)
    for i, col in enumerate(hist_cols):
        ws[f"{col}{r}"] = cd.ebit[i]
        style_input(ws[f"{col}{r}"], FMT_CURRENCY)
    r += 1

    #determin margin by dividing ebit by revenue
    ebit_margin_row = r
    ws[f"A{r}"] = "  EBIT Margin"
    style_label(ws[f"A{r}"], indent=1)
    for i, col in enumerate(hist_cols):
        ws[f"{col}{r}"] = f"={col}{ebit_row}/{col}{rev_row}"
        style_formula(ws[f"{col}{r}"], FMT_PCT)
    r += 1

    #hardcode depreciation and amortization
    da_row = r
    ws[f"A{r}"] = "D&A"
    style_label(ws[f"A{r}"], indent=1)
    for i, col in enumerate(hist_cols):
        if cd.da:
            ws[f"{col}{r}"] = cd.da[i]
        else:
            ws[f"{col}{r}"] = 0
        style_input(ws[f"{col}{r}"], FMT_CURRENCY)
    r += 1

    #hardcode capital expenditure
    capex_row = r
    ws[f"A{r}"] = "CapEx"
    style_label(ws[f"A{r}"], indent=1)
    for i, col in enumerate(hist_cols):
        ws[f"{col}{r}"] = cd.capex[i] if cd.capex else 0
        style_input(ws[f"{col}{r}"], FMT_CURRENCY)
    r += 1

    #hardcode net working capital
    nwc_row = r
    ws[f"A{r}"] = "Net Working Capital"
    style_label(ws[f"A{r}"], indent=1)
    for i, col in enumerate(hist_cols):
        ws[f"{col}{r}"] = cd.working_capital[i] if cd.working_capital else 0
        style_input(ws[f"{col}{r}"], FMT_CURRENCY)
    r += 1

    #calculate change in net working capital from current and prior nwc
    delta_nwc_hist_row = r
    ws[f"A{r}"] = "  Δ NWC"
    style_label(ws[f"A{r}"], indent=1)
    ws[f"{hist_cols[0]}{r}"] = "n/a"
    ws[f"{hist_cols[0]}{r}"].alignment = Alignment(horizontal="right")
    ws[f"{hist_cols[0]}{r}"].font = Font(name=FONT_NAME, color="595959", italic=True)
    for i in range(1, n_hist):
        ws[f"{hist_cols[i]}{r}"] = f"={hist_cols[i]}{nwc_row}-{hist_cols[i-1]}{nwc_row}"
        style_formula(ws[f"{hist_cols[i]}{r}"], FMT_CURRENCY)
    r += 1

    # ---- Projections ----
    r += 1
    ws[f"A{r}"] = "PROJECTIONS"
    ws[f"A{r}"].font = Font(name=FONT_NAME, bold=True)
    r += 1

    #pull driver rows from Assumptions sheet
    ar = refs["proj_rows"]

    #projected revenue
    proj_rev_row = r
    ws[f"A{r}"] = "Revenue"
    style_label(ws[f"A{r}"], indent=1)
    last_hist_col = hist_cols[-1]
    for i, col in enumerate(proj_cols):
        #Reference growth rate from Assumptions sheet (row ar['growth'], col B+i)
        a_col = get_column_letter(2 + i)
        #if it is the first year, multiply last historical year's revenue by (1 + Year 1 growth rate)
        if i == 0:
            ws[f"{col}{r}"] = f"={last_hist_col}{rev_row}*(1+Assumptions!{a_col}{ar['growth']})"
        #else, multiply prior preojection year's revenue by (1 + current growth rate)
        else:
            prev = proj_cols[i-1]
            ws[f"{col}{r}"] = f"={prev}{r}*(1+Assumptions!{a_col}{ar['growth']})"
        style_formula(ws[f"{col}{r}"], FMT_CURRENCY, link=True)
    r += 1
    
    ws[f"A{r}"] = "  Revenue Growth"
    style_label(ws[f"A{r}"], indent=1)
    for i, col in enumerate(proj_cols):
        a_col = get_column_letter(2 + i)
        ws[f"{col}{r}"] = f"=Assumptions!{a_col}{ar['growth']}"
        style_formula(ws[f"{col}{r}"], FMT_PCT, link=True)
    r += 1

    #projected ebit calculated as projected revenue * projected ebit margin
    proj_ebit_row = r
    ws[f"A{r}"] = "EBIT"
    style_label(ws[f"A{r}"], indent=1)
    for i, col in enumerate(proj_cols):
        a_col = get_column_letter(2 + i)
        ws[f"{col}{r}"] = f"={col}{proj_rev_row}*Assumptions!{a_col}{ar['ebit_margin']}"
        style_formula(ws[f"{col}{r}"], FMT_CURRENCY)
    r += 1

    ws[f"A{r}"] = "  EBIT Margin"
    style_label(ws[f"A{r}"], indent=1)
    for i, col in enumerate(proj_cols):
        a_col = get_column_letter(2 + i)
        ws[f"{col}{r}"] = f"=Assumptions!{a_col}{ar['ebit_margin']}"
        style_formula(ws[f"{col}{r}"], FMT_PCT, link=True)
    r += 1

    #net operating profit after tax calculated as EBIT*(1- tax rate)
    nopat_row = r
    ws[f"A{r}"] = "NOPAT = EBIT × (1 - t)"
    style_label(ws[f"A{r}"], indent=1)
    for i, col in enumerate(proj_cols):
        ws[f"{col}{r}"] = f"={col}{proj_ebit_row}*(1-{refs['tax_rate']})"
        style_formula(ws[f"{col}{r}"], FMT_CURRENCY, link=True)
    r += 1

    #d&a calculated as projected revenue multiplied by assumed d&a ratio
    proj_da_row = r
    ws[f"A{r}"] = "+ D&A"
    style_label(ws[f"A{r}"], indent=1)
    for i, col in enumerate(proj_cols):
        a_col = get_column_letter(2 + i)
        ws[f"{col}{r}"] = f"={col}{proj_rev_row}*Assumptions!{a_col}{ar['da_pct']}"
        style_formula(ws[f"{col}{r}"], FMT_CURRENCY)
    r += 1

    #capex calculated as projected revenue multiplied by assumed capex ratio
    proj_capex_row = r
    ws[f"A{r}"] = "− CapEx"
    style_label(ws[f"A{r}"], indent=1)
    for i, col in enumerate(proj_cols):
        a_col = get_column_letter(2 + i)
        ws[f"{col}{r}"] = f"={col}{proj_rev_row}*Assumptions!{a_col}{ar['capex_pct']}"
        style_formula(ws[f"{col}{r}"], FMT_CURRENCY)
    r += 1

    #net working capital as incremented revenue multiplied by nwc ratio
    proj_dnwc_row = r
    ws[f"A{r}"] = "− Δ NWC"
    style_label(ws[f"A{r}"], indent=1)
    for i, col in enumerate(proj_cols):
        a_col = get_column_letter(2 + i)
        prev_rev = hist_cols[-1] if i == 0 else proj_cols[i-1]
        ws[f"{col}{r}"] = (
            f"=({col}{proj_rev_row}-{prev_rev}{proj_rev_row if i>0 else rev_row})"
            f"*Assumptions!{a_col}{ar['wc_pct']}"
        )
        style_formula(ws[f"{col}{r}"], FMT_CURRENCY)
    r += 1

    #finalize free cash flow as NOPAT + D&A - CapEx - change in net working capital
    fcf_row = r
    ws[f"A{r}"] = "Free Cash Flow"
    ws[f"A{r}"].font = Font(name=FONT_NAME, bold=True)
    ws[f"A{r}"].border = BORDER_THIN
    for i, col in enumerate(proj_cols):
        ws[f"{col}{r}"] = (
            f"={col}{nopat_row}+{col}{proj_da_row}-{col}{proj_capex_row}-{col}{proj_dnwc_row}"
        )
        style_result(ws[f"{col}{r}"], FMT_CURRENCY)
    r += 2

    #Discounting
    period_row = r
    ws[f"A{r}"] = "Discount Period (years)"
    style_label(ws[f"A{r}"], indent=1)
    for i, col in enumerate(proj_cols):
        ws[f"{col}{r}"] = i + 1
        style_formula(ws[f"{col}{r}"], FMT_NUM)
    r += 1

    df_row = r
    ws[f"A{r}"] = "Discount Factor = 1/(1+WACC)^t"
    style_label(ws[f"A{r}"], indent=1)
    for i, col in enumerate(proj_cols):
        ws[f"{col}{r}"] = f"=1/(1+{refs['wacc']})^{col}{period_row}"
        style_formula(ws[f"{col}{r}"], "0.0000", link=True)
    r += 1

    #multiply each years free cash flow by discount factor to get present value of that year's fcf
    pv_fcf_row = r
    ws[f"A{r}"] = "PV of FCF"
    style_label(ws[f"A{r}"], indent=1, bold=True)
    for i, col in enumerate(proj_cols):
        ws[f"{col}{r}"] = f"={col}{fcf_row}*{col}{df_row}"
        style_result(ws[f"{col}{r}"], FMT_CURRENCY)
    r += 2

    # ---- Terminal Value ----
    ws[f"A{r}"] = "TERMINAL VALUE (Gordon Growth)"
    ws[f"A{r}"].font = Font(name=FONT_NAME, bold=True)
    r += 1

    #calculate terminal value as FCF_projected_years * (1+g)/(WACC-g)
    tv_row = r
    last_proj_col = proj_cols[-1]
    ws[f"A{r}"] = "Terminal Value = FCF₅ × (1+g) / (WACC - g)"
    style_label(ws[f"A{r}"], indent=1)
    ws[f"{last_proj_col}{r}"] = (
        f"={last_proj_col}{fcf_row}*(1+{refs['g']})/({refs['wacc']}-{refs['g']})"
    )
    style_result(ws[f"{last_proj_col}{r}"], FMT_CURRENCY)
    r += 1

    #present value of terminal value as Terminal Value * Projected Years discount factor
    pv_tv_row = r
    ws[f"A{r}"] = "PV of Terminal Value"
    style_label(ws[f"A{r}"], indent=1, bold=True)
    ws[f"{last_proj_col}{r}"] = f"={last_proj_col}{tv_row}*{last_proj_col}{df_row}"
    style_result(ws[f"{last_proj_col}{r}"], FMT_CURRENCY)
    r += 2

    # ---- Valuation Summary on this sheet ----
    ws[f"A{r}"] = "ENTERPRISE & EQUITY VALUE"
    ws[f"A{r}"].font = Font(name=FONT_NAME, bold=True)
    r += 1

    #add up the sum of all present values
    sum_pv_row = r
    ws[f"A{r}"] = "Sum of PV of Projected FCFs"
    style_label(ws[f"A{r}"], indent=1)
    first_pv = f"{proj_cols[0]}{pv_fcf_row}"
    last_pv = f"{proj_cols[-1]}{pv_fcf_row}"
    ws[f"B{r}"] = f"=SUM({first_pv}:{last_pv})"
    style_formula(ws[f"B{r}"], FMT_CURRENCY)
    r += 1

    pv_tv_summary_row = r
    ws[f"A{r}"] = "PV of Terminal Value"
    style_label(ws[f"A{r}"], indent=1)
    ws[f"B{r}"] = f"={last_proj_col}{pv_tv_row}"
    style_formula(ws[f"B{r}"], FMT_CURRENCY)
    r += 1

    ev_row = r
    ws[f"A{r}"] = "Enterprise Value"
    style_label(ws[f"A{r}"], indent=1, bold=True)
    ws[f"B{r}"] = f"=B{sum_pv_row}+B{pv_tv_summary_row}"
    style_result(ws[f"B{r}"], FMT_CURRENCY)
    r += 1

    nd_row = r
    ws[f"A{r}"] = "− Net Debt"
    style_label(ws[f"A{r}"], indent=1)
    ws[f"B{r}"] = f"={refs['net_debt']}"
    style_formula(ws[f"B{r}"], FMT_CURRENCY, link=True)
    r += 1

    eq_row = r
    ws[f"A{r}"] = "Equity Value"
    style_label(ws[f"A{r}"], indent=1, bold=True)
    ws[f"B{r}"] = f"=B{ev_row}-B{nd_row}"
    style_result(ws[f"B{r}"], FMT_CURRENCY)
    r += 1

    sh_row = r
    ws[f"A{r}"] = "÷ Shares Outstanding (mm)"
    style_label(ws[f"A{r}"], indent=1)
    ws[f"B{r}"] = f"={refs['shares_out']}"
    style_formula(ws[f"B{r}"], FMT_NUM, link=True)
    r += 1

    iv_row = r
    ws[f"A{r}"] = "Implied Share Price"
    style_label(ws[f"A{r}"], indent=1, bold=True)
    ws[f"B{r}"] = f"=B{eq_row}/B{sh_row}"
    style_result(ws[f"B{r}"], FMT_CURRENCY_DEC)
    r += 1

    #save key cell references for Summary sheet
    refs["enterprise_value"] = f"'FCF Projections'!$B${ev_row}"
    refs["equity_value"] = f"'FCF Projections'!$B${eq_row}"
    refs["implied_price"] = f"'FCF Projections'!$B${iv_row}"
    refs["sum_pv_fcf"] = f"'FCF Projections'!$B${sum_pv_row}"
    refs["pv_tv"] = f"'FCF Projections'!$B${pv_tv_summary_row}"

    # Column widths
    widths = {"A": 36}
    for c in range(2, 2 + n_hist + n_proj):
        widths[get_column_letter(c)] = 14
    set_col_widths(ws, widths)
    return refs


"""
Sensitivity table: implied share price across grid of WACC (cols) × g (rows).

Recomputes valuation from scratch for each cell, using the SAME projected FCFs
that live on the FCF sheet
"""
def build_sensitivity_sheet(wb: Workbook, cd: CompanyData, refs: dict):

    ws = wb.create_sheet("Sensitivity Table")
    ws.merge_cells("A1:K1")
    ws["A1"] = "Sensitivity Analysis: Implied Share Price vs. WACC and Terminal Growth"
    style_header(ws["A1"], large=True)
    ws.row_dimensions[1].height = 26

    #Show base case
    ws["A3"] = "Base Case WACC:"
    style_label(ws["A3"], bold=True)
    ws["B3"] = f"={refs['wacc']}"
    style_formula(ws["B3"], FMT_PCT, link=True)

    ws["A4"] = "Base Case Terminal Growth:"
    style_label(ws["A4"], bold=True)
    ws["B4"] = f"={refs['g']}"
    style_formula(ws["B4"], FMT_PCT, link=True)

    ws["A5"] = "Base Case Implied Price:"
    style_label(ws["A5"], bold=True)
    ws["B5"] = f"={refs['implied_price']}"
    style_result(ws["B5"], FMT_CURRENCY_DEC)

    # ---- Build the sensitivity table ----
    #Columns are the WACC values
    #rows are the g values
    n_wacc = 7
    n_g = 7
    wacc_step = 0.005
    g_step = 0.0025

    table_top = 8
    ws.cell(row=table_top, column=1, value="Implied Share Price ($)")
    style_subheader(ws.cell(row=table_top, column=1))
    ws.cell(row=table_top, column=2, value="WACC →")
    style_subheader(ws.cell(row=table_top, column=2))

    #fill WACC values across the top row
    wacc_header_row = table_top
    for j in range(n_wacc):
        col = get_column_letter(3 + j)
        #offset from base: -3, -2, -1, 0, +1, +2, +3 steps
        offset = (j - (n_wacc // 2)) * wacc_step
        ws[f"{col}{wacc_header_row}"] = f"={refs['wacc']}+{offset}"
        style_subheader(ws[f"{col}{wacc_header_row}"])
        ws[f"{col}{wacc_header_row}"].number_format = FMT_PCT

    #Row label for g
    ws.cell(row=table_top + 1, column=1, value="↓ Terminal Growth (g)")
    ws.cell(row=table_top + 1, column=1).font = Font(name=FONT_NAME, bold=True, italic=True)
    ws.cell(row=table_top + 1, column=1).alignment = Alignment(horizontal="left")

    #need column letters for the projected FCF row on FCF sheet
    #re-derive the FCF sheet "Free Cash Flow" cells.
    n_hist = len(cd.fiscal_years)
    n_proj = PROJECTION_YEARS
    proj_col_letters = [get_column_letter(2 + n_hist + i) for i in range(n_proj)]
    fcf_sheet = wb["FCF Projections"]

    #walk down column A to find "free cash flow"
    fcf_row = None
    for r_idx in range(1, fcf_sheet.max_row + 1):
        if fcf_sheet.cell(row=r_idx, column=1).value == "Free Cash Flow":
            fcf_row = r_idx
            break
    if fcf_row is None:
        raise RuntimeError("Could not locate FCF row on FCF Projections sheet")

    #implied share price = (Σ FCFt/(1+w)^t  +  FCF5*(1+g)/(w-g)/(1+w)^5  − Net Debt) / Shares
    sum_terms = []
    for t, c in enumerate(proj_col_letters, start=1):
        # discount of period-t FCF using cell-level w
        sum_terms.append(f"'FCF Projections'!{c}{fcf_row}/(1+$WACC$)^{t}")
    pv_fcf_expr = "+".join(sum_terms)

    last_c = proj_col_letters[-1]
    tv_expr = f"'FCF Projections'!{last_c}{fcf_row}*(1+$G$)/($WACC$-$G$)/(1+$WACC$)^{n_proj}"
    eq_expr = f"({pv_fcf_expr}+{tv_expr}-{refs['net_debt']})/{refs['shares_out']}"

    #place the formulas via replace $WACC$ and $G$ tokens with cell refs
    g_col = "B"
    for i in range(n_g):
        r = table_top + 1 + i
        # g value down the side (in column B)
        offset = (i - (n_g // 2)) * g_step
        ws[f"{g_col}{r}"] = f"={refs['g']}+{offset}"
        style_subheader(ws[f"{g_col}{r}"])
        ws[f"{g_col}{r}"].number_format = FMT_PCT

        for j in range(n_wacc):
            cell = ws.cell(row=r, column=3 + j)
            wacc_cell = f"{get_column_letter(3 + j)}${wacc_header_row}"
            g_cell = f"$B{r}"
            formula = "=" + eq_expr.replace("$WACC$", wacc_cell).replace("$G$", g_cell)
            cell.value = formula
            cell.number_format = FMT_CURRENCY_DEC
            cell.font = Font(name=FONT_NAME, color="000000")
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="right")
            #highlight the base case
            if i == n_g // 2 and j == n_wacc // 2:
                cell.fill = RESULT_FILL
                cell.font = Font(name=FONT_NAME, color="000000", bold=True)

    # ---- Upside/downside table ----
    ud_top = table_top + n_g + 3
    ws.cell(row=ud_top, column=1, value="Upside / (Downside) vs. Current Price")
    style_subheader(ws.cell(row=ud_top, column=1))
    ws.cell(row=ud_top, column=2, value="WACC →")
    style_subheader(ws.cell(row=ud_top, column=2))

    for j in range(n_wacc):
        col = get_column_letter(3 + j)
        ws[f"{col}{ud_top}"] = f"={col}{wacc_header_row}"
        style_subheader(ws[f"{col}{ud_top}"])
        ws[f"{col}{ud_top}"].number_format = FMT_PCT

    for i in range(n_g):
        r_price = table_top + 1 + i
        r_ud = ud_top + 1 + i
        ws[f"{g_col}{r_ud}"] = f"={g_col}{r_price}"
        style_subheader(ws[f"{g_col}{r_ud}"])
        ws[f"{g_col}{r_ud}"].number_format = FMT_PCT
        for j in range(n_wacc):
            c = get_column_letter(3 + j)
            cell = ws.cell(row=r_ud, column=3 + j)
            cell.value = f"={c}{r_price}/{refs['current_price']}-1"
            cell.number_format = FMT_PCT
            cell.font = Font(name=FONT_NAME)
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="right")
            if i == n_g // 2 and j == n_wacc // 2:
                cell.fill = RESULT_FILL
                cell.font = Font(name=FONT_NAME, bold=True)

    # Add conditional formatting (color scale) on both tables
    from openpyxl.formatting.rule import ColorScaleRule
    cs_rule = ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    )
    price_rng = (f"{get_column_letter(3)}{table_top+1}:"
                 f"{get_column_letter(2+n_wacc)}{table_top+n_g}")
    ws.conditional_formatting.add(price_rng, cs_rule)

    ud_rng = (f"{get_column_letter(3)}{ud_top+1}:"
              f"{get_column_letter(2+n_wacc)}{ud_top+n_g}")
    ws.conditional_formatting.add(ud_rng, cs_rule)

    widths = {"A": 32, "B": 14}
    for c in range(3, 3 + n_wacc):
        widths[get_column_letter(c)] = 14
    set_col_widths(ws, widths)


def build_summary_sheet(wb: Workbook, cd: CompanyData, refs: dict):
    #insert as first sheet
    ws = wb.create_sheet("Summary", 0)
    ws.merge_cells("A1:D1")
    ws["A1"] = f"DCF Valuation Summary - {cd.name} ({cd.ticker})"
    style_header(ws["A1"], large=True)
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"Generated from {cd.data_source}.   "\
               f"Historical period: {cd.fiscal_years[0]}-{cd.fiscal_years[-1]}.   "\
               f"Projection horizon: {PROJECTION_YEARS} years."
    ws["A2"].font = Font(name=FONT_NAME, italic=True, color="595959", size=10)
    ws.merge_cells("A2:D2")

    rows = [
        ("Company Info", "header"),
        ("Ticker", f"={refs.get('_ticker', '')}", "@", "label_only"),  # we'll patch
        ("Sector", cd.sector, "@", "label_only"),
        ("Reporting Currency", cd.currency, "@", "label_only"),
        ("", "blank"),
        ("Market Snapshot", "header"),
        ("Current Share Price", f"={refs['current_price']}", FMT_CURRENCY_DEC, "link"),
        ("Shares Outstanding (mm)", f"={refs['shares_out']}", FMT_NUM, "link"),
        ("Market Cap ($mm)", f"={refs['market_cap']}", FMT_CURRENCY, "link"),
        ("Total Debt ($mm)", f"={refs['total_debt']}", FMT_CURRENCY, "link"),
        ("Cash & ST Investments ($mm)", f"={refs['cash']}", FMT_CURRENCY, "link"),
        ("Net Debt ($mm)", f"={refs['net_debt']}", FMT_CURRENCY, "link"),
        ("", "blank"),
        ("Cost of Capital", "header"),
        ("WACC", f"={refs['wacc']}", FMT_PCT, "link"),
        ("Terminal Growth Rate", f"={refs['g']}", FMT_PCT, "link"),
        ("", "blank"),
        ("DCF Valuation Output", "header"),
        ("Sum of PV of FCFs ($mm)", f"={refs['sum_pv_fcf']}", FMT_CURRENCY, "link"),
        ("PV of Terminal Value ($mm)", f"={refs['pv_tv']}", FMT_CURRENCY, "link"),
        ("Enterprise Value ($mm)", f"={refs['enterprise_value']}", FMT_CURRENCY, "result"),
        ("Less: Net Debt ($mm)", f"={refs['net_debt']}", FMT_CURRENCY, "link"),
        ("Equity Value ($mm)", f"={refs['equity_value']}", FMT_CURRENCY, "result"),
        ("÷ Shares Outstanding (mm)", f"={refs['shares_out']}", FMT_NUM, "link"),
        ("Implied Intrinsic Share Price", f"={refs['implied_price']}", FMT_CURRENCY_DEC, "result"),
        ("Current Market Price", f"={refs['current_price']}", FMT_CURRENCY_DEC, "link"),
        ("Upside / (Downside)", None, FMT_PCT, "upside"),
        ("", "blank"),
        ("Recommendation", None, "@", "reco"),
    ]

    r = 4
    upside_row = None
    implied_row = None
    market_row = None
    for entry in rows:
        if entry[1] == "header":
            ws.merge_cells(f"A{r}:D{r}")
            ws[f"A{r}"] = entry[0]
            style_subheader(ws[f"A{r}"])
            ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center", indent=0)
            r += 1
            continue
        if entry[1] == "blank":
            r += 1
            continue

        label, value, fmt, kind = entry
        ws[f"A{r}"] = label
        style_label(ws[f"A{r}"], indent=1, bold=(kind == "result"))

        if kind == "label_only":
            ws[f"B{r}"] = value if not str(value).startswith("=") else value
            ws[f"B{r}"].font = Font(name=FONT_NAME)
            ws[f"B{r}"].border = BORDER_THIN
            ws[f"B{r}"].number_format = fmt or "@"
            ws[f"B{r}"].alignment = Alignment(horizontal="left")
        elif kind == "link":
            ws[f"B{r}"] = value
            style_formula(ws[f"B{r}"], fmt, link=True)
        elif kind == "result":
            ws[f"B{r}"] = value
            style_result(ws[f"B{r}"], fmt)
        elif kind == "upside":
            upside_row = r
            ws[f"B{r}"] = f"=B{implied_row}/B{market_row}-1"
            style_result(ws[f"B{r}"], fmt)
        elif kind == "reco":
            ws[f"B{r}"] = (
                f'=IF(B{upside_row}>0.2,"BUY - significantly undervalued",'
                f'IF(B{upside_row}>0.05,"ACCUMULATE - modestly undervalued",'
                f'IF(B{upside_row}>-0.05,"HOLD - fairly valued",'
                f'IF(B{upside_row}>-0.2,"REDUCE - modestly overvalued",'
                f'"SELL - significantly overvalued"))))'
            )
            ws[f"B{r}"].font = Font(name=FONT_NAME, bold=True, size=12)
            ws[f"B{r}"].fill = RESULT_FILL
            ws[f"B{r}"].border = BORDER_THIN
            ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[r].height = 22

        #remember key rows
        if "Implied Intrinsic" in label:
            implied_row = r
        if "Current Market Price" in label:
            market_row = r

        r += 1

    # Patch the "Ticker" row (we couldn't easily inline it above)
    for r_idx in range(1, ws.max_row + 1):
        if ws.cell(row=r_idx, column=1).value == "Ticker":
            ws.cell(row=r_idx, column=2, value=cd.ticker)
            ws.cell(row=r_idx, column=2).font = Font(name=FONT_NAME)
            ws.cell(row=r_idx, column=2).border = BORDER_THIN
            ws.cell(row=r_idx, column=2).alignment = Alignment(horizontal="left")
            break

    set_col_widths(ws, {"A": 34, "B": 22, "C": 4, "D": 4})


# ----------------------------------------------------------------------------
# MAIN
# ----------------------------------------------------------------------------
def main():
    #read the users input ticker
    ticker = TICKER
    if len(sys.argv) > 1:
        ticker = sys.argv[1].upper()

    print(f"=== DCF Valuation Model: {ticker} ===\n")

    #1. Fetch
    cd = fetch_company_data(ticker)

    #2 Build workbook
    print("[2/4] Building Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    refs = {}
    refs.update(build_assumptions_sheet(wb, cd))
    refs = build_wacc_sheet(wb, cd, refs)
    refs = build_fcf_sheet(wb, cd, refs)
    build_sensitivity_sheet(wb, cd, refs)
    build_summary_sheet(wb, cd, refs)

    #order the sheets
    desired = ["Summary", "Assumptions", "FCF Projections", "WACC", "Sensitivity Table"]
    wb._sheets = [wb[name] for name in desired if name in wb.sheetnames]

    out_path = OUTPUT_FILE
    wb.save(out_path)
    print(f"      Saved {out_path}")

    #3. recalculate formulas via LibreOffice
    import os
    recalc_script = "/mnt/skills/public/xlsx/scripts/recalc.py"
    recalc_ran = False
    if os.path.exists(recalc_script):
        print("[3/4] Recalculating formulas...")
        import subprocess
        try:
            recalc = subprocess.run(
                ["python3", recalc_script, out_path],
                capture_output=True, text=True, timeout=120,
            )
            print(recalc.stdout[-500:] if recalc.stdout else "")
            if recalc.returncode != 0:
                print("      [warn] recalc had non-zero exit:", recalc.stderr[-300:])
            else:
                recalc_ran = True
        except Exception as e:
            print(f"      [warn] recalc failed ({type(e).__name__}: {e})")
    else:
        print("[3/4] Skipping formula recalc (helper not present on this OS).")
        print("      Excel will recalculate every formula automatically when you open the file.")

    #4. Verify by reading back values 
    if recalc_ran:
        print("[4/4] Verifying values...")
        from openpyxl import load_workbook
        wb_chk = load_workbook(out_path, data_only=True)
        s = wb_chk["Summary"]
        for r_idx in range(1, s.max_row + 1):
            label = s.cell(row=r_idx, column=1).value
            val = s.cell(row=r_idx, column=2).value
            if label in ("WACC", "Implied Intrinsic Share Price",
                         "Current Market Price", "Upside / (Downside)",
                         "Enterprise Value ($mm)", "Equity Value ($mm)",
                         "Recommendation"):
                print(f"      {label}: {val}")
    else:
        print("[4/4] Skipping value verification - open the file in Excel to see results.")

    print(f"\nDone. Output: {out_path}")


if __name__ == "__main__":
    main()